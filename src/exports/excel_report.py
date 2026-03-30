"""
Excel Report Exporter — Generates polished .xlsx workbook with KPIs, pivot table, and charts.

Usage:
    python -m src.exports.excel_report
    python -m src.exports.excel_report --output reports/south_shore_report.xlsx

Requires: pip install openpyxl
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.utils.constants import PHASES, TARGET_EMOTIONS
from src.utils.db import get_connection
from src.utils.logger import log

HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
KPI_FONT = Font(bold=True, size=14, color="FF4500")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    """Apply header styling to a row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _load_data() -> pd.DataFrame:
    """Load analyzed posts from DuckDB."""
    conn = get_connection()
    try:
        df = conn.execute("SELECT * FROM posts_full").fetchdf()
    except Exception:
        df = conn.execute("""
            SELECT c.*, e.vader_compound, e.sentiment_label,
                   e.emo_fear, e.emo_anger, e.emo_sadness, e.emo_joy,
                   e.emo_surprise, e.emo_disgust, e.emo_gratitude, e.emo_pride,
                   e.dominant_emotion, t.topic_label
            FROM posts_clean c
            LEFT JOIN posts_emotions e ON c.id = e.id
            LEFT JOIN posts_topics t ON c.id = t.id
        """).fetchdf()
    conn.close()
    return df


def _create_kpi_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 1: Executive KPIs."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    kpis = [
        ("South Shore Sentiment Study", ""),
        ("", ""),
        ("KPI", "Value"),
        ("Total Posts Analyzed", f"{len(df):,}"),
        (
            "Date Range",
            f"{df['dt_utc'].min():%Y-%m-%d} → {df['dt_utc'].max():%Y-%m-%d}"
            if "dt_utc" in df
            else "N/A",
        ),
        ("Platforms", f"{df['platform'].nunique()} (Reddit, YouTube, News)"),
        (
            "Avg Sentiment (VADER)",
            f"{df['vader_compound'].mean():.3f}" if "vader_compound" in df else "N/A",
        ),
        (
            "% Negative Posts",
            f"{(df['sentiment_label'] == 'negative').mean() * 100:.1f}%"
            if "sentiment_label" in df
            else "N/A",
        ),
        (
            "% Positive Posts",
            f"{(df['sentiment_label'] == 'positive').mean() * 100:.1f}%"
            if "sentiment_label" in df
            else "N/A",
        ),
        (
            "Dominant Emotion",
            df["dominant_emotion"].mode().iloc[0]
            if "dominant_emotion" in df and not df["dominant_emotion"].mode().empty
            else "N/A",
        ),
        ("Temporal Phases", str(len(PHASES))),
        ("Analysis Window", "88 days (Sep 16 – Dec 12, 2025)"),
    ]

    for row_idx, (label, value) in enumerate(kpis, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    # Style title
    ws.cell(1, 1).font = Font(bold=True, size=16, color="FF4500")
    # Style header row
    ws.cell(3, 1).font = HEADER_FONT
    ws.cell(3, 1).fill = HEADER_FILL
    ws.cell(3, 2).font = HEADER_FONT
    ws.cell(3, 2).fill = HEADER_FILL
    # Style KPI values
    for r in range(4, len(kpis) + 1):
        ws.cell(r, 2).font = Font(bold=True, size=12)
        ws.cell(r, 1).border = THIN_BORDER
        ws.cell(r, 2).border = THIN_BORDER

    log.info("Created Executive Summary sheet")


def _create_pivot_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 2: Phase × Platform × Emotion pivot table."""
    ws = wb.create_sheet("Phase × Platform Pivot")

    if "phase" not in df.columns or "platform" not in df.columns:
        ws.cell(1, 1, "Insufficient data for pivot")
        return

    emo_cols = [f"emo_{e}" for e in TARGET_EMOTIONS if f"emo_{e}" in df.columns]
    pivot = (
        df.groupby(["phase", "platform"])[["vader_compound"] + emo_cols]
        .mean()
        .round(3)
        .reset_index()
    )

    # Sort by phase order
    phase_order = {k: i for i, k in enumerate(PHASES.keys())}
    pivot["_order"] = pivot["phase"].map(phase_order)
    pivot = pivot.sort_values(["_order", "platform"]).drop(columns=["_order"])

    for r_idx, row in enumerate(dataframe_to_rows(pivot, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _style_header_row(ws, len(pivot.columns))

    # Auto-width columns
    for col_idx in range(1, len(pivot.columns) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = 16

    log.info(f"Created pivot sheet: {len(pivot)} rows")


def _create_chart_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 3: Platform comparison bar chart."""
    ws = wb.create_sheet("Platform Chart")

    if "platform" not in df.columns or "vader_compound" not in df.columns:
        ws.cell(1, 1, "Insufficient data for chart")
        return

    platform_stats = (
        df.groupby("platform")["vader_compound"].agg(["mean", "count"]).round(3).reset_index()
    )
    platform_stats.columns = ["Platform", "Avg Sentiment", "Post Count"]

    for r_idx, row in enumerate(dataframe_to_rows(platform_stats, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _style_header_row(ws, 3)

    # Create bar chart
    chart = BarChart()
    chart.title = "Average Sentiment by Platform"
    chart.y_axis.title = "VADER Compound Score"
    chart.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(platform_stats) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(platform_stats) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    ws.add_chart(chart, "E2")
    log.info("Created platform chart sheet")


def _create_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Sheet 4: Raw data sample (first 500 rows) for transparency."""
    ws = wb.create_sheet("Raw Data (Sample)")

    keep_cols = [
        c
        for c in [
            "id",
            "platform",
            "source",
            "dt_utc",
            "phase",
            "word_count",
            "vader_compound",
            "sentiment_label",
            "dominant_emotion",
            "emo_fear",
            "emo_anger",
            "emo_sadness",
            "emo_joy",
            "emo_gratitude",
            "topic_label",
        ]
        if c in df.columns
    ]

    # sample = df[keep_cols].head(500)
    sample = df[keep_cols].head(500).copy()

    # Strip timezone info — Excel doesn't support timezone-aware datetimes
    for col in sample.columns:
        if pd.api.types.is_datetime64_any_dtype(sample[col]):
            sample[col] = sample[col].dt.tz_localize(None)

    for r_idx, row in enumerate(dataframe_to_rows(sample, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:
                cell.border = THIN_BORDER

    _style_header_row(ws, len(keep_cols))

    for col_idx in range(1, len(keep_cols) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = 18

    log.info(f"Created raw data sheet: {len(sample)} rows")


def generate_excel_report(output_path: str = "reports/south_shore_report.xlsx") -> str:
    """Generate the complete Excel workbook."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    log.info("Loading data from DuckDB...")
    df = _load_data()
    log.info(f"Loaded {len(df)} posts")

    wb = Workbook()
    _create_kpi_sheet(wb, df)
    _create_pivot_sheet(wb, df)
    _create_chart_sheet(wb, df)
    _create_raw_data_sheet(wb, df)

    wb.save(str(output))
    log.info(f"✅ Excel report saved: {output}")
    return str(output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel report")
    parser.add_argument("--output", default="reports/south_shore_report.xlsx")
    args = parser.parse_args()
    generate_excel_report(args.output)
